import psycopg2
import datetime
from openpyxl import Workbook,load_workbook

db = psycopg2.connect(user = "postgres",
                      password = "1",
                      host = "localhost",
                      port = "5432",
                      database = "ANT")
crs=db.cursor()

wb = load_workbook("ERP.xlsx",  read_only=True)

ws = wb.sheetnames

for name in ws:
    columnNamesArr=[]
    columnNames=''
    header_cells_generator=wb[name].iter_rows(max_row=1)
    for header_cells_tuple in header_cells_generator:
        for i in range(len(header_cells_tuple)):
            columnNamesArr.append('"'+header_cells_tuple[i].value+'"')
    columnNames=', '.join(columnNamesArr)

    for row in range(2,wb[name].max_row+1):
        insertValuesArr=[]
        insertValues=''
        for column in range(1,wb[name].max_column+1):
            if type(wb[name].cell(row,column).value)==str:
                insertValuesArr.append("'"+wb[name].cell(row,column).value+"'")
            else:
                if wb[name].cell(row,column).value == None:
                    insertValuesArr.append('null')
                elif type(wb[name].cell(row,column).value)== datetime.datetime:                   
                    insertValuesArr.append('TO_TIMESTAMP('+wb[name].cell(row,column).value.strftime("'%m-%d-%Y %H:%M:%S'")+", 'DD-MM-YYYY HH24:MI:SS')")
                else:
                    insertValuesArr.append(wb[name].cell(row,column).value)
                #print(type(wb[name].cell(row,column).value))
        insertValues=', '.join([str(elem) for elem in insertValuesArr])
        insertCommand='INSERT INTO "PUBLIC"."'+ name +'" ('+ columnNames +') VALUES('+insertValues+');'

        try:
            crs.execute(insertCommand)
            db.commit()
            f= open("_SuccessLog"+name+".txt","a")
            f.write("\n"+insertCommand)
            f.close()
        except:
            f= open("_ErrorLog"+name+".txt","a")
            f.write("\n"+insertCommand)
            f.close()

print("success")
